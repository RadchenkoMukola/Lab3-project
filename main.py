from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import PySimpleGUI as sg


def input_cells():
    return [
        [sg.Text(item, size=size5, pad=(0, 0)),
         sg.Input("",size=size2, enable_events=True, key=item, expand_x=True)]
            for item in input_items]


def fk_cells():
    return [
        [sg.Text(item, size=size2, pad=(0, 0)),
         sg.Input("",size = size3, enable_events=True, key=item, expand_x=True)]
            for item in fk_items]


def sn_cells():
    return [
        [sg.Text(item, size=size1, pad=(0, 0)),
         sg.Input("",size = size3, enable_events=True, key=item, expand_x=True)]
            for item in sn_items]


def cat_cells():
 return [
  [sg.Text(item, size=5, pad=(0, 0)),
   sg.Input("", size=size3, enable_events=True, key=item, expand_x=True)]
  for item in cat_items]


def gem_cells():
 return [
  [sg.Text(item, size=18, pad=(0, 0)),
   sg.Input("", size=3, enable_events=True, key=item, expand_x=True)]
  for item in gem_items]


def at_cells():
 return [
  [sg.Text(item, size=10, pad=(0, 0)),
   sg.Input("", size=size3, enable_events=True, key=item, expand_x=True)]
  for item in at_items]


def post_cells():
 return [
  [sg.Text(item, size=18, pad=(0, 0)),
   sg.Input("", size=size3, enable_events=True, key=item, expand_x=True)]
  for item in post_items]


def eco_cells():
 return [
  [sg.Text(item, size=20, pad=(0, 0)),
   sg.Input("", size=15, enable_events=True, key=item, expand_x=True)]
  for item in eco_items]


def kat_cells():
 return [
  [sg.Text(item, size=25, pad=(0, 0)),
   sg.Input("", size=3, enable_events=True, key=item, expand_x=True)]
  for item in kat_items]


def gas_cells():
 return [
  [sg.Text(item, size=15, pad=(0, 0)),
   sg.Input("", size=5, enable_events=True, key=item, expand_x=True)]
  for item in gas_items]


def sin_cells():
 return [
  [sg.Text(item, size=15, pad=(0, 0)),
   sg.Input("", size=3, enable_events=True, key=item, expand_x=True)]
  for item in sin_items]


def kt_cells():
 return [
  [sg.Text(item, size=size2, pad=(0, 0)),
   sg.Input("", size=15, enable_events=True, key=item, expand_x=True)]
  for item in kt_items]


def lik_cells():
 return [
  [sg.Text(item, size=35, pad=(0, 0)),
   sg.Input("", size=3, enable_events=True, key=item, expand_x=True)]
  for item in lik_items]


def prm_cells():
 return [
  [sg.Text(item, size=50, pad=(0, 0)),
   sg.Input("", size=30, enable_events=True, key=item, expand_x=True)]
  for item in prm_items]

size1,size2,size3,size4,size5 = 40,10,2,12,32

input_items  = ('Рік', 'Номер','Госпіталізовані -1 чи амбулаторні -0','№ ІХ','ПІБ','Пацієнт первинний -1, повторний -0','Дата народження','Вік','Місце проживання','Телефон','Лікар','Стать ч-1 ж-0','Зріст(см)','Вага п','Вага в','Скарги, міс','Діагностована ЛГ,міс')
fk_items = ('ЛГ ','1.1.','1.2.','1.3.','1.4.1.','1.4.2.','1.4.3.','1.4.4.','1.4.5.','1’','2','3','4.1.','4.2.','5','І ст.','ІІ ст.','ІІІ ст.','I ФК ВОЗ','IІ ФК ВОЗ','IІІ ФК ВОЗ','IV ФК ВОЗ',)
sn_items = ('СН І','СН ІІА','СН ІІБ','ФП','ТП','Асцит','Гідроперикард','Плеврит, гідроторакс','ХОЗЛ','СОАС','Індекс AHI','Мінімальна сатурація','ГПМК','Анемія','Залізодефіцитний стан','Гіпотиреоз 1, гіпертиреоз 2, еутиреоз 3','ВВС','Після операції з приводу вади','ДМПП','ДМШП','ВАП','ІНШІ ВВС','ЗКИД ЛП 0, ПЛ 1, перехрестний 2','Тромбоутворення в н/кінцівках 1-є, 0- відсутне','Кровотечі','Синкопе')
cat_items = ('САТ п','ДАТ п','ЧСС п','САТ в','ДАТ в','ЧСС в')
gem_items = ('Гемоглобін','Еритроцити','Гематокрит','MCV','MCH','Тромбоцити','Лейкоцити','ШОЄ','МНО','Заг.хол.','ТГ','К.','білірубін','креатинін','кліренс креатиніну','Сечова кислота','АЛТ ','АСТ','глюкоза','HBsAg','Anti-HCV','ВИЧ','Ntpro-BNP (ДІЛА)','Ntpro-BNP (Інститут)','ТТГ','Феритин','Anti-ENA скринінг','Результат:0-нег, 1-поз')
at_items =('Цент АТ','СРПХ м','СРПХ е','R-CAVI','L-CAVI','R-ABI','L-ABI')
post_items = ('6ХХ м пост.','6ХХ бали','SpO2 до','SpO2 після','6ХХ м вип.','6ХХ бали.','SpO2 до.','SpO2 після.')
eco_items = ('ЕхоКс','Аорта','ЛП','S ЛП','Інд.ЛП','S ПП','Інд.ПП','МШП','ЗСЛШ','КДР ЛШ','КСР ЛШ','КДО ЛШ','КСО ЛШ','УО ЛШ','ФВ','ММ ЛШ','E/A МК','DT МК','БС ПШ/ЛШ','Стінка ПШ','S ПШ діаст','S ПШ сист','Фракц.скор.ПШ','Позд.розПШ','Попер.розПШ','S’','TAPSE мм','Швидкість рег. на ТСК ','Т АСС ','ЛА','НПВ','Колаб{1>50,25<(0,5)<50,0<25}','Сист.ТЛА','Сер.ТЛА','ІЕ діастола','ІЕ систола','АК нед-ть','МК нед-ть','ТК нед-ть','ЛК нед-ть','ЙЛГ 0-низ, 1-сер, 2-вис','Висновок')
kat_items = ('КПС Інс-1,поза-2,не роб-0','Дата кат','Сист.Т.ЛА','Діас.Т.ЛА','сер ТЛА','Сист.Т.ПШ ПШ','Діас.Т.ПШ','сер ТПШ','Сист.Т. ПП','Діас.Т.ПП','сер ПП','ТЗЛА','ХОК','Серц.індекс','Удар.викид','Удар.індекс','PVR','PVR Wood','ЧСС','САТ','ДАТ','Сер АТ','SVR','TPR','ГДД','LVSW ','LVSWI','RVSW','RVSWI','Вазореак.тест','СТЛА п т','ДТЛА п т','Сер тиск ЛА п т','ТЗЛА.','ХОК.','Серц.інд.','Удар.викид.','Удар.інд.','PVR.','ЧСС.','САТ.','ДАТ.','сер АТ.','Результат.','Проба з піднятими н/к','СТЛА.','ДТЛА.','СерТЛА.','ТЗЛА..','Кров ЛА..')
gas_items = ('pH','pO2','pCO2','Hct','Na','K','Ca','THb','HCO3','SvO2','BE в','Р50 смеш вен','Кров артер','pH ','pO2 ','pCO2 ','Hct ','Na ','K ','Ca ','THb ','HCO3 ','SaO2 ','BE в ','Р50 арт ','Оцінка ризику')
sin_items = ('Спирометрія','ФЖЕЛ','ФЖЕЛ%','ОФВ1','ОФВ1%','ОФВ1/ФЖЕЛ','ОФВ1/ФЖЕЛ%','ЖЕЛ','ЖЕЛ%','ОФВ1/ЖЕЛ','ОФВ1/ЖЕЛ%','Хв спож кисню','ДО','ЛН 1-рес,2-обс','Висновок ','БодіПГ','TLC','TLC%','ДифЗдЛегень','DLCO','DLCO%')
kt_items = ('КТ, МРТ ОГК','Дата','Висновок.')
lik_items = ('Лік:1-д,1`-н,2-а,3-сіл,4-т,5-в,6-б,7-амб,8-р,9-с','Фурос.в/в','Фурос.в/м','Дофамін','Сілденафіл','ДозаS','Тадалафіл','ДозаTAD','Вентавіс','ДозаVEN','АК 1-а, 2-д, 3-н, 4-вер','ДозаAK','Фуросемід','ДозаF','Ант. Альд.  1 Верошпірон, 2 -еплеренон','ДозаAA','Сер.глік.','ДозаSG','Небіволол 1, карведілол 2','ДозаBB','ЕРА:Б-1,а-2','ДозаERA','ОАК:вар-1, кс-2, сінк-3,прад-4','ДозаOAK','Тріфас','ДозаT','Кордарон','ДозаK','АСК','Залізо 1 в/м, 2 per os, 3 в/в','Оксигенотерапія','Оптимальна Терапія')
prm_items = ('Примітки','ТЕА 0-рек,1-пров,2-відм.БАП 3-рек,4-пров, 5-відм 6-не рек','Дата теа','Рекомендации Кулика','Декомпенсація','Дата дек','Трансплантація 1-пров, 0 -рек','Дата транс','Смерть','Дата см')
sg.theme('SandyBeach')
sg.SetOptions(font = ('Arial', 10, 'bold'))

layout = [
    [sg.Frame('Основні дані',input_cells(),vertical_alignment='top', expand_y=True, expand_x=True),sg.Column(fk_cells(),vertical_scroll_only=True,scrollable=True,  vertical_alignment='top', expand_y=True, expand_x=True),sg.Column(sn_cells(),vertical_scroll_only=True,scrollable=True,  vertical_alignment='top', expand_y=True, expand_x=True),sg.Column(gem_cells(),vertical_scroll_only=True,scrollable=True,  vertical_alignment='top', expand_y=True, expand_x=True),sg.Column(eco_cells(),vertical_scroll_only=True,scrollable=True,vertical_alignment='top', expand_y=True, expand_x=True),sg.Column(kat_cells(),vertical_scroll_only=True,scrollable=True,vertical_alignment='top', expand_y=True, expand_x=True),sg.Column(gas_cells(),vertical_scroll_only=True,scrollable=True,vertical_alignment='top', expand_y=True, expand_x=True)]
 ,[sg.Column(cat_cells(), vertical_scroll_only=True,scrollable=True,  vertical_alignment='top', expand_y=True, expand_x=True),sg.Column(at_cells(),vertical_scroll_only=True,scrollable=True,  vertical_alignment='top', expand_y=True, expand_x=True),sg.Column(post_cells(),vertical_scroll_only=True,scrollable=True,vertical_alignment='top', expand_y=True, expand_x=True),sg.Column(sin_cells(),vertical_scroll_only=True,scrollable=True,vertical_alignment='top', expand_y=True, expand_x=True),sg.Column(kt_cells(), vertical_scroll_only=True,scrollable=True,  vertical_alignment='top', expand_y=True, expand_x=True),sg.Column(lik_cells(), vertical_scroll_only=True,scrollable=True,  vertical_alignment='top', expand_y=True, expand_x=True),sg.Column(prm_cells(), vertical_scroll_only=True,scrollable=True,  vertical_alignment='top', expand_y=True, expand_x=True),[sg.Button('Додати',button_color='green', key='calcA'),sg.Button('Очистити',button_color='blue', key='clear')]]]


window = sg.Window('Exelfiller', layout,resizable=True)


while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED:
        break
    elif event == 'calcA':

     for i in values:
      if str(values[i]) == "0,5":
       values[i] = 0.5

     #1 'Рік', 'Номер','Госпіталізовані -1 чи амбулаторні -0','№ ІХ','ПІБ','Пацієнт первинний -1, повторний -0','Дата народження','Вік','Місце проживання','Телефон','Лікар','Стать ч-1 ж-0','Зріст(см)','Вага п','Вага в','Скарги, міс','Діагностована ЛГ,міс'
     Rik = None if str((values['Рік'])) == "" else float(values['Рік'])
     Number = None if str((values['Номер'])) == "" else float(values['Номер'])
     Gos = None if str((values['Госпіталізовані -1 чи амбулаторні -0'])) == ""  else float(values['Госпіталізовані -1 чи амбулаторні -0'])
     NoIX =None if str((values['№ ІХ'])) == "" else float(values['№ ІХ'])
     PIB = str(values['ПІБ'])
     Per = None if str((values['Пацієнт первинний -1, повторний -0'])) == "" else float(values['Пацієнт первинний -1, повторний -0'])
     Birth = str(values['Дата народження'])
     Vik = None if str((values['Вік'])) == "" else float(values['Вік'])
     Mis = str(values['Місце проживання'])
     Pho = str(values['Телефон'])
     Doc = str(values['Лікар'])
     Sex = None if str((values['Стать ч-1 ж-0'])) == "" else  float(values['Стать ч-1 ж-0'])
     weightp = None if str((values['Вага п'])) == "" else float(values['Вага п'])
     weightv = None if str((values['Вага в'])) == "" else float(values['Вага в'])
     height = None if str((values['Зріст(см)'])) == "" else float(values['Зріст(см)'])
     Scarm = None if str((values['Скарги, міс'])) == "" else float(values['Скарги, міс'])
     Diag = None if str((values['Діагностована ЛГ,міс'])) == "" else float(values['Діагностована ЛГ,міс'])
     #2 'ЛГ ','1.1.','1.2.','1.3.','1.4.1.','1.4.2.','1.4.3.','1.4.4.','1.4.5.','1’','2','3','4.1.','4.2.','5','І ст.','ІІ ст.','ІІІ ст.','I ФК ВОЗ','IІ ФК ВОЗ','IІІ ФК ВОЗ','IV ФК ВОЗ'

     Lg = None if str((values['ЛГ '])) == "" else float(values['ЛГ '])
     N11 = None if str((values['1.1.'])) == "" else float(values['1.1.'])
     N12 = None if str((values['1.2.'])) == "" else float(values['1.2.'])
     N13 = None if str((values['1.3.'])) == "" else float(values['1.3.'])
     N141 = None if str((values['1.4.1.'])) == "" else float(values['1.4.1.'])
     N142 = None if str((values['1.4.2.'])) == "" else float(values['1.4.2.'])
     N143 = None if str((values['1.4.3.'])) == "" else float(values['1.4.3.'])
     N144 = None if str((values['1.4.4.'])) == "" else float(values['1.4.4.'])
     N145 = None if str((values['1.4.5.'])) == "" else float(values['1.4.5.'])
     N1 = None if str((values['1’'])) == "" else float(values['1’'])
     N2 = None if str((values['2'])) == "" else float(values['2'])
     N3 = None if str((values['3'])) == "" else float(values['3'])
     N41 = None if str((values['4.1.'])) == "" else float(values['4.1.'])
     N42 = None if str((values['4.2.'])) == "" else float(values['4.2.'])
     N5 = None if str((values['5'])) == "" else float(values['5'])
     St1 = None if str((values['І ст.'])) == "" else float(values['І ст.'])
     St2 = None if str((values['ІІ ст.'])) == "" else float(values['ІІ ст.'])
     St3 = None if str((values['ІІІ ст.'])) == "" else float(values['ІІІ ст.'])
     Fk1 = None if str((values['I ФК ВОЗ'])) == "" else float(values['I ФК ВОЗ'])
     Fk2 = None if str((values['IІ ФК ВОЗ'])) == "" else float(values['IІ ФК ВОЗ'])
     Fk3 = None if str((values['IІІ ФК ВОЗ'])) == "" else float(values['IІІ ФК ВОЗ'])
     Fk4 = None if str((values['IV ФК ВОЗ'])) == "" else float(values['IV ФК ВОЗ'])

     # 3 'СН І','СН ІІА','СН ІІБ','ФП','ТП','Асцит','Гідроперикард','Плеврит, гідроторакс','ХОЗЛ','СОАС','Індекс AHI','Мінімальна сатурація','ГПМК','Анемія','Залізодефіцитний стан','Гіпотиреоз 1, гіпертиреоз 2, еутиреоз 3','ВВС','Після операції з приводу вади','ДМПП','ДМШП','ВАП','ІНШІ ВВС','ЗКИД ЛП 0, ПЛ 1, перехрестний 2','Тромбоутворення в н/кінцівках 1-є, 0- відсутне','Кровотечі','Синкопе'
     Sn1 = None if str((values['СН І'])) == "" else float(values['СН І'])
     Sn2a = None if str((values['СН ІІА'])) == "" else float(values['СН ІІА'])
     Sn2b = None if str((values['СН ІІБ'])) == "" else float(values['СН ІІБ'])
     Fp = None if str((values['ФП'])) == "" else float(values['ФП'])
     Tp = None if str((values['ТП'])) == "" else float(values['ТП'])
     As = None if str((values['Асцит'])) == "" else float(values['Асцит'])
     Gip = None if str((values['Гідроперикард'])) == "" else float(values['Гідроперикард'])
     Plv = None if str((values['Плеврит, гідроторакс'])) == "" else float(values['Плеврит, гідроторакс'])
     Hoz = None if str((values['ХОЗЛ'])) == "" else float(values['ХОЗЛ'])
     Soa = None if str((values['СОАС'])) == "" else float(values['СОАС'])
     ANI = None if str((values['Індекс AHI'])) == "" else float(values['Індекс AHI'])
     Mins = None if str((values['Мінімальна сатурація'])) == "" else float(values['Мінімальна сатурація'])
     Gpmk = None if str((values['ГПМК'])) == "" else float(values['ГПМК'])
     Anm = None if str((values['Анемія'])) == "" else float(values['Анемія'])
     Zald = None if str((values['Залізодефіцитний стан'])) == "" else float(values['Залізодефіцитний стан'])
     Gp = None if str((values['Гіпотиреоз 1, гіпертиреоз 2, еутиреоз 3'])) == "" else float(values['Гіпотиреоз 1, гіпертиреоз 2, еутиреоз 3'])
     BBC = None if str((values['ВВС'])) == "" else float(values['ВВС'])
     Pis = None if str((values['Після операції з приводу вади'])) == "" else float(values['Після операції з приводу вади'])
     DMP = None if str((values['ДМПП'])) == "" else float(values['ДМПП'])
     DSP = None if str((values['ДМШП'])) == "" else float(values['ДМШП'])
     Vap = None if str((values['ВАП'])) == "" else float(values['ВАП'])
     IBBC = None if str((values['ІНШІ ВВС'])) == "" else float(values['ІНШІ ВВС'])
     Zkd = None if str((values['ЗКИД ЛП 0, ПЛ 1, перехрестний 2'])) == "" else float(values['ЗКИД ЛП 0, ПЛ 1, перехрестний 2'])
     Tmb = None if str((values['Тромбоутворення в н/кінцівках 1-є, 0- відсутне'])) == "" else float(values['Тромбоутворення в н/кінцівках 1-є, 0- відсутне'])
     Krv = None if str((values['Кровотечі'])) == "" else float(values['Кровотечі'])
     Sin = None if str((values['Синкопе'])) == "" else float(values['Синкопе'])


     #4 'САТ п','ДАТ п','ЧСС п','САТ в','ДАТ в','ЧСС в'
     Catp = None if str((values['САТ п'])) == "" else float(values['САТ п'])
     Datp = None if str((values['ДАТ п'])) == "" else float(values['ДАТ п'])
     Csp = None if str((values['ЧСС п'])) == "" else float(values['ЧСС п'])
     Catv = None if str((values['САТ в'])) == "" else float(values['САТ в'])
     Datv = None if str((values['ДАТ в'])) == "" else float(values['ДАТ в'])
     Csv = None if str((values['ЧСС в'])) == "" else float(values['ЧСС в'])

     #5'Гемоглобін','Еритроцити','Гематокрит','MCV','MCH','Тромбоцити','Лейкоцити','ШОЄ','МНО','Заг.хол.','ТГ','К.','білірубін','креатинін','кліренс креатиніну','Сечова кислота','АЛТ ','АСТ','глюкоза','HBsAg','Anti-HCV','ВИЧ','Ntpro-BNP (ДІЛА)','Ntpro-BNP (Інститут)','ТТГ','Феритин','Anti-ENA скринінг','Результат: 0- негативний, 1 - позитивний'
     Gem = None if str((values['Гемоглобін'])) == "" else float(values['Гемоглобін'])
     Ert = None if str((values['Еритроцити'])) == "" else float(values['Еритроцити'])
     Gemkr =None if str((values['Гематокрит'])) == "" else float(values['Гематокрит'])
     MCV = None if str((values['MCV'])) == "" else float(values['MCV'])
     MCH = None if str((values['MCH'])) == "" else float(values['MCH'])
     Trobm = None if str((values['Тромбоцити'])) == "" else float(values['Тромбоцити'])
     Lek = None if str((values['Лейкоцити'])) == "" else float(values['Лейкоцити'])
     Shoe = None if str((values['ШОЄ'])) == "" else float(values['ШОЄ'])
     MNO = None if str((values['МНО'])) == "" else float(values['МНО'])
     Zag = None if str((values['Заг.хол.'])) == "" else float(values['Заг.хол.'])
     Tg = None if str((values['ТГ'])) == "" else float(values['ТГ'])
     K1 = None if str((values['К.'])) == "" else float(values['К.'])
     Bil = None if str((values['білірубін'])) == "" else float(values['білірубін'])
     Cre = None if str((values['креатинін'])) == "" else float(values['креатинін'])
     Cli = None if str((values['кліренс креатиніну'])) == "" else float(values['кліренс креатиніну'])
     Sech = None if str((values['Сечова кислота'])) == "" else float(values['Сечова кислота'])
     Alt = None if str((values['АЛТ '])) == "" else float(values['АЛТ '])
     Ast = None if str((values['АСТ'])) == "" else float(values['АСТ'])
     Glu = None if str((values['глюкоза'])) == "" else float(values['глюкоза'])
     HBsAg = None if str((values['HBsAg'])) == "" else float(values['HBsAg'])
     AHCV = None if str((values['Anti-HCV'])) == "" else float(values['Anti-HCV'])
     Vich = None if str((values['ВИЧ'])) == "" else float(values['ВИЧ'])
     Ntpro1 = None if str((values['Ntpro-BNP (ДІЛА)'])) == "" else float(values['Ntpro-BNP (ДІЛА)'])
     Ntpro2 = None if str((values['Ntpro-BNP (Інститут)'])) == "" else float(values['Ntpro-BNP (Інститут)'])
     TTG = None if str((values['ТТГ'])) == "" else float(values['ТТГ'])
     Fer = None if str((values['Феритин'])) == "" else float(values['Феритин'])
     AENA = None if str((values['Anti-ENA скринінг'])) == "" else float(values['Anti-ENA скринінг'])
     Rez = None if str((values['Результат:0-нег, 1-поз'])) == "" else float(values['Результат:0-нег, 1-поз'])

     #6'Цент АТ','СРПХ м','СРПХ е','R-CAVI','L-CAVI','R-ABI','L-ABI'
     CAT = None if str((values['Цент АТ'])) == "" else float(values['Цент АТ'])
     CRPXM = None if str((values['СРПХ м'])) == "" else float(values['СРПХ м'])
     CRPXE = None if str((values['СРПХ е'])) == "" else float(values['СРПХ е'])
     RCav = None if str((values['R-CAVI'])) == "" else float(values['R-CAVI'])
     LCav = None if str((values['L-CAVI'])) == "" else float(values['L-CAVI'])
     RABI = None if str((values['R-ABI'])) == "" else float(values['R-ABI'])
     LABI = None if str((values['L-ABI'])) == "" else float(values['L-ABI'])



     #7'6ХХ м пост.','6ХХ бали','SpO2 до','SpO2 після','6ХХ м вип.','6ХХ бали.','SpO2 до.','SpO2 після.'
     xxp1 = None if str((values['6ХХ м пост.'])) == "" else float(values['6ХХ м пост.'])
     xxb1 = None if str((values['6ХХ бали'])) == "" else float(values['6ХХ бали'])
     Spo1 = None if str((values['SpO2 до'])) == "" else float(values['SpO2 до'])
     Spop1 = None if str((values['SpO2 після'])) == "" else float(values['SpO2 після'])
     xxp2 = None if str((values['6ХХ м вип.'])) == "" else float(values['6ХХ м вип.'])
     xxb2 = None if str((values['6ХХ бали.'])) == "" else float(values['6ХХ бали.'])
     Spo2 = None if str((values['SpO2 до.'])) == "" else float(values['SpO2 до.'])
     Spop2 = None if str((values['SpO2 після.'])) == "" else float(values['SpO2 після.'])


     #8'ЕхоКс','Аорта','ЛП','S ЛП','Інд.ЛП','S ПП','Інд.ПП','МШП','ЗСЛШ','КДР ЛШ','КСР ЛШ','КДО ЛШ','КСО ЛШ','УО ЛШ','ФВ','ММ ЛШ','E/A МК','DT МК','БС ПШ/ЛШ','Стінка ПШ','S ПШ діаст','S ПШ сист','Фракц.скор.ПШ','Позд.розПШ','Попер.розПШ','S’','TAPSE мм','Швидкість рег. на ТСК ','Т АСС ','ЛА','НПВ','Колаб{1>50,25<(0,5)<50,0<25}','Сист.ТЛА','Сер.ТЛА','ІЕ діастола','ІЕ систола','АК нед-ть','МК нед-ть','ТК нед-ть','ЛК нед-ть','ЙЛГ 0-низ, 1-сер, 2-вис','Висновок'
     EXOKs = None if str((values['ЕхоКс'])) == "" else float(values['ЕхоКс'])
     AORT = None if str((values['Аорта'])) == "" else float(values['Аорта'])
     LP = None if str((values['ЛП'])) == "" else float(values['ЛП'])
     SLP = None if str((values['S ПП'])) == "" else float(values['S ПП'])
     INDLP = None if str((values['Інд.ЛП'])) == "" else float(values['Інд.ЛП'])
     SPP = None if str((values['S ЛП'])) == "" else float(values['S ЛП'])
     INDPP = None if str((values['Інд.ПП'])) == "" else float(values['Інд.ПП'])
     MSP = None if str((values['МШП'])) == "" else float(values['МШП'])
     ZSLS = None if str((values['ЗСЛШ'])) == "" else float(values['ЗСЛШ'])
     KDRLS = None if str((values['КДР ЛШ'])) == "" else float(values['КДР ЛШ'])
     KSRLS = None if str((values['КСР ЛШ'])) == "" else float(values['КСР ЛШ'])
     KDOLS = None if str((values['КДО ЛШ'])) == "" else float(values['КДО ЛШ'])
     KSOLS = None if str((values['КСО ЛШ'])) == "" else float(values['КСО ЛШ'])
     YOLS = None if str((values['УО ЛШ'])) == "" else float(values['УО ЛШ'])
     FV = None if str((values['ФВ'])) == "" else float(values['ФВ'])
     MMLS = None if str((values['ММ ЛШ'])) == "" else float(values['ММ ЛШ'])
     EAMK = None if str((values['E/A МК'])) == "" else float(values['E/A МК'])
     DTMK = None if str((values['DT МК'])) == "" else float(values['DT МК'])
     BSPS = None if str((values['БС ПШ/ЛШ'])) == "" else float(values['БС ПШ/ЛШ'])
     SNPS = None if str((values['Стінка ПШ'])) == "" else float(values['Стінка ПШ'])
     SPSD = None if str((values['S ПШ діаст'])) == "" else float(values['S ПШ діаст'])
     SPSS = None if str((values['S ПШ сист'])) == "" else float(values['S ПШ сист'])
     FSPS = None if str((values['Фракц.скор.ПШ'])) == "" else float(values['Фракц.скор.ПШ'])
     POZD = None if str((values['Позд.розПШ'])) == "" else float(values['Позд.розПШ'])
     POPR = None if str((values['Попер.розПШ'])) == "" else float(values['Попер.розПШ'])
     Ss = None if str((values['S’'])) == "" else float(values['S’'])
     TAPSE = None if str((values['TAPSE мм'])) == "" else float(values['TAPSE мм'])
     SRNT = None if str((values['Швидкість рег. на ТСК '])) == "" else float(values['Швидкість рег. на ТСК '])
     TASS = None if str((values['Т АСС '])) == "" else float(values['Т АСС '])
     LA = None if str((values['ЛА'])) == "" else float(values['ЛА'])
     NPV = None if str((values['НПВ'])) == "" else float(values['НПВ'])
     KOLB = None if str((values['Колаб{1>50,25<(0,5)<50,0<25}'])) == "" else float(values['Колаб{1>50,25<(0,5)<50,0<25}'])
     SIstl = None if str((values['Сист.ТЛА'])) == "" else float(values['Сист.ТЛА'])
     Sertl = None if str((values['Сер.ТЛА'])) == "" else float(values['Сер.ТЛА'])
     IED = None if str((values['ІЕ діастола'])) == "" else float(values['ІЕ діастола'])
     IES = None if str((values['ІЕ систола'])) == "" else float(values['ІЕ систола'])
     AKN = None if str((values['АК нед-ть'])) == "" else float(values['АК нед-ть'])
     MKN = None if str((values['МК нед-ть'])) == "" else float(values['МК нед-ть'])
     TKN = None if str((values['ТК нед-ть'])) == "" else float(values['ТК нед-ть'])
     LKN = None if str((values['ЛК нед-ть'])) == "" else float(values['ЛК нед-ть'])
     YLG = None if str((values['ЙЛГ 0-низ, 1-сер, 2-вис'])) == "" else float(values['ЙЛГ 0-низ, 1-сер, 2-вис'])
     VYS = None if str((values['Висновок'])) == "" else str(values['Висновок'])

     #9 'КПС Інс-1,поза-2,не роб-0','Дата кат','Сист.Т.ЛА','Діас.Т.ЛА','сер ТЛА','Сист.Т.ПШ ПШ','Діас.Т.ПШ','сер ТПШ','Сист.Т. ПП','Діас.Т.ПП','сер ПП','ТЗЛА','ХОК','Серц.індекс','Удар.викид','Удар.індекс','PVR','PVR Wood','ЧСС','САТ','ДАТ','Сер АТ','SVR','TPR','ГДД','LVSW ','LVSWI','RVSW','RVSWI','Вазореак.тест','СТЛА п т','ДТЛА п т','Сер тиск ЛА п т','ТЗЛА.','ХОК.','Серц.інд.','Удар.викид.','Удар.інд.','PVR.','ЧСС.','САТ.','ДАТ.','сер АТ.','Результат.','Проба з піднятими н/к','СТЛА.','ДТЛА.','СерТЛА.','ТЗЛА..','Кров ЛА..'
     KPS = None if str((values['КПС Інс-1,поза-2,не роб-0'])) == "" else float(values['КПС Інс-1,поза-2,не роб-0'])
     DATAKAT = None if str((values['Дата кат'])) == "" else str(values['Дата кат'])
     STL1 = None if str((values['Сист.Т.ЛА'])) == "" else float(values['Сист.Т.ЛА'])
     DTL1 = None if str((values['Діас.Т.ЛА'])) == "" else float(values['Діас.Т.ЛА'])
     SerTLA1 = None if str((values['сер ТЛА'])) == "" else float(values['сер ТЛА'])
     SSTLPS = None if str((values['Сист.Т.ПШ ПШ'])) == "" else float(values['Сист.Т.ПШ ПШ'])
     DSTLPS = None if str((values['Діас.Т.ПШ'])) == "" else float(values['Діас.Т.ПШ'])
     SerTPS = None if str((values['сер ТПШ'])) == "" else float(values['сер ТПШ'])
     SSTLPP = None if str((values['Сист.Т. ПП'])) == "" else float(values['Сист.Т. ПП'])
     DSTLPP = None if str((values['Діас.Т.ПП'])) == "" else float(values['Діас.Т.ПП'])
     SerPP = None if str((values['сер ПП'])) == "" else float(values['сер ПП'])
     TZLA1 = None if str((values['ТЗЛА'])) == "" else float(values['ТЗЛА'])
     HOK1 = None if str((values['ХОК'])) == "" else float(values['ХОК'])
     SERI1 = None if str((values['Серц.індекс'])) == "" else float(values['Серц.індекс'])
     UV1 = None if str((values['Удар.викид'])) == "" else float(values['Удар.викид'])
     UI1 = None if str((values['Удар.індекс'])) == "" else float(values['Удар.індекс'])
     PVR1 = None if str((values['PVR'])) == "" else float(values['PVR'])
     PVRw = None if str((values['PVR Wood'])) == "" else float(values['PVR Wood'])
     CSS1 = None if str((values['ЧСС'])) == "" else float(values['ЧСС'])
     SAT1 = None if str((values['САТ'])) == "" else float(values['САТ'])
     DAT1 = None if str((values['ДАТ'])) == "" else float(values['ДАТ'])
     SerAT1 = None if str((values['Сер АТ'])) == "" else float(values['Сер АТ'])
     SVR = None if str((values['SVR'])) == "" else float(values['SVR'])
     TPR = None if str((values['TPR'])) == "" else float(values['TPR'])
     GDD = None if str((values['ГДД'])) == "" else float(values['ГДД'])
     LVSW = None if str((values['LVSW '])) == "" else float(values['LVSW '])
     LVSWI = None if str((values['LVSWI'])) == "" else float(values['LVSWI'])
     RVSW = None if str((values['RVSW'])) == "" else float(values['RVSW'])
     RVSWI = None if str((values['RVSWI'])) == "" else float(values['RVSWI'])
     VRT = None if str((values['Вазореак.тест'])) == "" else float(values['Вазореак.тест'])
     SSTLA2 = None if str((values['СТЛА п т'])) == "" else float(values['СТЛА п т'])
     DSTLA2 = None if str((values['ДТЛА п т'])) == "" else float(values['ДТЛА п т'])
     SerTLA2 = None if str((values['Сер тиск ЛА п т'])) == "" else float(values['Сер тиск ЛА п т'])
     TZLA2 = None if str((values['ТЗЛА.'])) == "" else float(values['ТЗЛА.'])
     HOK2 = None if str((values['ХОК.'])) == "" else float(values['ХОК.'])
     SERI2 = None if str((values['Серц.інд.'])) == "" else float(values['Серц.інд.'])
     UV2 = None if str((values['Удар.викид.'])) == "" else float(values['Удар.викид.'])
     UI2 = None if str((values['Удар.інд.'])) == "" else float(values['Удар.інд.'])
     PVR2 = None if str((values['PVR.'])) == "" else float(values['PVR.'])
     CSS2 = None if str((values['ЧСС.'])) == "" else float(values['ЧСС.'])
     SAT2 = None if str((values['САТ.'])) == "" else float(values['САТ.'])
     DAT2 = None if str((values['ДАТ.'])) == "" else float(values['ДАТ.'])
     SerAT2 = None if str((values['сер АТ.'])) == "" else float(values['сер АТ.'])
     Rez2 = None if str((values['Результат.'])) == "" else str(values['Результат.'])
     PROB = None if str((values['Проба з піднятими н/к'])) == "" else float(values['Проба з піднятими н/к'])
     STLA = None if str((values['СТЛА.'])) == "" else float(values['СТЛА.'])
     DTLA = None if str((values['ДТЛА.'])) == "" else float(values['ДТЛА.'])
     SerTLA3 = None if str((values['СерТЛА.'])) == "" else float(values['СерТЛА.'])
     TZLA3 = None if str((values['ТЗЛА..'])) == "" else float(values['ТЗЛА..'])
     KROVLA = None if str((values['Кров ЛА..'])) == "" else float(values['Кров ЛА..'])

     #10'pH','pO2','pCO2','Hct','Na','K','Ca','THb','HCO3','SvO2','BE в','Р50 смеш вен','Кров артер','pH ','pO2 ','pCO2 ','Hct ','Na ','K ','Ca ','THb ','HCO3 ','SaO2 ','BE в ','Р50 арт ','Оцінка ризику'
     pH1 = None if str((values['pH'])) == "" else float(values['pH'])
     pO21 = None if str((values['pO2'])) == "" else float(values['pO2'])
     pCO21 = None if str((values['pCO2'])) == "" else float(values['pCO2'])
     Hct1 = None if str((values['Hct'])) == "" else float(values['Hct'])
     Na1 = None if str((values['Na'])) == "" else float(values['Na'])
     Ka = None if str((values['K'])) == "" else float(values['K'])
     Ca1 = None if str((values['Ca'])) == "" else float(values['Ca'])
     THb1 = None if str((values['THb'])) == "" else float(values['THb'])
     HCO31 = None if str((values['HCO3'])) == "" else float(values['HCO3'])
     SvO21 = None if str((values['SvO2'])) == "" else float(values['SvO2'])
     BE1 = None if str((values['BE в'])) == "" else float(values['BE в'])
     P50v = None if str((values['Р50 смеш вен'])) == "" else float(values['Р50 смеш вен'])
     Kra = None if str((values['Кров артер'])) == "" else float(values['Кров артер'])
     pH2 = None if str((values['pH '])) == "" else float(values['pH '])
     pO22 = None if str((values['pO2 '])) == "" else float(values['pO2 '])
     pCO22 = None if str((values['pCO2 '])) == "" else float(values['pCO2 '])
     Hct2  = None if str((values['Hct '])) == "" else float(values['Hct '])
     Na2 = None if str((values['Na '])) == "" else float(values['Na '])
     Kb = None if str((values['K '])) == "" else float(values['K '])
     Ca2 = None if str((values['Ca '])) == "" else float(values['Ca '])
     THb2 = None if str((values['THb '])) == "" else float(values['THb '])
     HCO32 = None if str((values['HCO3 '])) == "" else float(values['HCO3 '])
     SvO22 = None if str((values['SaO2 '])) == "" else float(values['SaO2 '])
     BE2 = None if str((values['BE в '])) == "" else float(values['BE в '])
     P50a = None if str((values['Р50 арт '])) == "" else float(values['Р50 арт '])
     Ryz = None if str((values['Оцінка ризику'])) == "" else float(values['Оцінка ризику'])

     #11'Спирометрія','ФЖЕЛ','ФЖЕЛ%','ОФВ1','ОФВ1%','ОФВ1/ФЖЕЛ','ОФВ1/ФЖЕЛ%','ЖЕЛ','ЖЕЛ%','ОФВ1/ЖЕЛ','ОФВ1/ЖЕЛ%','Хв спож кисню','ДО','ЛН 1-рес,2-обс','Висновок ','БодіПГ','TLC','TLC%','ДифЗдЛегень','DLCO','DLCO%'
     Spr = None if str((values['Спирометрія'])) == "" else float(values['Спирометрія'])
     Fgel = None if str((values['ФЖЕЛ'])) == "" else float(values['ФЖЕЛ'])
     Fgelp = None if str((values['ФЖЕЛ%'])) == "" else str(values['ФЖЕЛ%'])
     OFv = None if str((values['ОФВ1'])) == "" else float(values['ОФВ1'])
     OFvp = None if str((values['ОФВ1%'])) == "" else str(values['ОФВ1%'])
     OFFG = None if str((values['ОФВ1/ФЖЕЛ'])) == "" else float(values['ОФВ1/ФЖЕЛ'])
     OFFGp = None if str((values['ОФВ1/ФЖЕЛ%'])) == "" else str(values['ОФВ1/ФЖЕЛ%'])
     Gel = None if str((values['ЖЕЛ'])) == "" else float(values['ЖЕЛ'])
     Gelp = None if str((values['ЖЕЛ%'])) == "" else str(values['ЖЕЛ%'])
     OFG = None if str((values['ОФВ1/ЖЕЛ'])) == "" else float(values['ОФВ1/ЖЕЛ'])
     OFGp = None if str((values['ОФВ1/ЖЕЛ%'])) == "" else str(values['ОФВ1/ЖЕЛ%'])
     HSK = None if str((values['Хв спож кисню'])) == "" else float(values['Хв спож кисню'])
     Do = None if str((values['ДО'])) == "" else float(values['ДО'])
     Ln = None if str((values['ЛН 1-рес,2-обс'])) == "" else float(values['ЛН 1-рес,2-обс'])
     VYS2 = None if str((values['Висновок '])) == "" else str(values['Висновок '])
     Bodi = None if str((values['БодіПГ'])) == "" else float(values['БодіПГ'])
     TLS = None if str((values['TLC'])) == "" else float(values['TLC'])
     TLSp = None if str((values['TLC%'])) == "" else str(values['TLC%'])
     Dif = None if str((values['ДифЗдЛегень'])) == "" else float(values['ДифЗдЛегень'])
     DLCO = None if str((values['DLCO'])) == "" else float(values['DLCO'])
     DLCOp = None if str((values['DLCO%'])) == "" else str(values['DLCO%'])

     #12'КТ, МРТ ОГК','Дата','Висновок.'
     KTMO = None if str((values['КТ, МРТ ОГК'])) == "" else float(values['КТ, МРТ ОГК'])
     Datak = None if str((values['Дата'])) == "" else str(values['Дата'])
     VYS3 = None if str((values['Висновок.'])) == "" else str(values['Висновок.'])

     #13'Лік:1-д,1`-н,2-а,3-сіл,4-т,5-в,6-б,7-амб,8-р,9-с','Фурос.в/в','Фурос.в/м','Дофамін','Сілденафіл','ДозаS','Тадалафіл','ДозаTAD','Вентавіс','ДозаVEN','АК 1-а, 2-д, 3-н, 4-вер','ДозаAK','Фуросемід','ДозаF','Ант. Альд.  1 Верошпірон, 2 -еплеренон','ДозаAA','Сер.глік.','ДозаSG','Небіволол 1, карведілол 2','ДозаBB','ЕРА:Б-1,а-2','ДозаERA','ОАК:вар-1, кс-2, сінк-3,прад-4','ДозаOAK','Тріфас','ДозаT','Кордарон','ДозаK','АСК','Залізо 1 в/м, 2 per os, 3 в/в','Оксигенотерапія','Оптимальна Терапія'
     Lik = None if str((values['Лік:1-д,1`-н,2-а,3-сіл,4-т,5-в,6-б,7-амб,8-р,9-с'])) == "" else float(values['Лік:1-д,1`-н,2-а,3-сіл,4-т,5-в,6-б,7-амб,8-р,9-с'])
     Furv = None if str((values['Фурос.в/в'])) == "" else float(values['Фурос.в/в'])
     Furm = None if str((values['Фурос.в/м'])) == "" else float(values['Фурос.в/м'])
     Dof = None if str((values['Дофамін'])) == "" else float(values['Дофамін'])
     Sli = None if str((values['Сілденафіл'])) == "" else float(values['Сілденафіл'])
     DOZS = None if str((values['ДозаS'])) == "" else float(values['ДозаS'])
     TAD = None if str((values['Тадалафіл'])) == "" else float(values['Тадалафіл'])
     DOZT = None if str((values['ДозаTAD'])) == "" else float(values['ДозаTAD'])
     VEN = None if str((values['Вентавіс'])) == "" else float(values['Вентавіс'])
     DOZV = None if str((values['ДозаVEN'])) == "" else float(values['ДозаVEN'])
     AK = None if str((values['АК 1-а, 2-д, 3-н, 4-вер'])) == "" else float(values['АК 1-а, 2-д, 3-н, 4-вер'])
     DOZAK = None if str((values['ДозаAK'])) == "" else float(values['ДозаAK'])
     Fyr = None if str((values['Фуросемід'])) == "" else float(values['Фуросемід'])
     DOZF = None if str((values['ДозаF'])) == "" else float(values['ДозаF'])
     AA = None if str((values['Ант. Альд.  1 Верошпірон, 2 -еплеренон'])) == "" else float(values['Ант. Альд.  1 Верошпірон, 2 -еплеренон'])
     DOZAA = None if str((values['ДозаAA'])) == "" else float(values['ДозаAA'])
     SERG = None if str((values['Сер.глік.'])) == "" else float(values['Сер.глік.'])
     DOZSG = None if str((values['ДозаSG'])) == "" else float(values['ДозаSG'])
     VV = None if str((values['Небіволол 1, карведілол 2'])) == "" else float(values['Небіволол 1, карведілол 2'])
     DOZVV = None if str((values['ДозаBB'])) == "" else float(values['ДозаBB'])
     ERA = None if str((values['ЕРА:Б-1,а-2'])) == "" else float(values['ЕРА:Б-1,а-2'])
     DOZERA = None if str((values['ДозаERA'])) == "" else float(values['ДозаERA'])
     OAK = None if str((values['ОАК:вар-1, кс-2, сінк-3,прад-4'])) == "" else float(values['ОАК:вар-1, кс-2, сінк-3,прад-4'])
     DOZOAK = None if str((values['ДозаOAK'])) == "" else float(values['ДозаOAK'])
     TRI = None if str((values['Тріфас'])) == "" else float(values['Тріфас'])
     DOZTRI = None if str((values['ДозаT'])) == "" else float(values['ДозаT'])
     KOR = None if str((values['Кордарон'])) == "" else float(values['Кордарон'])
     DOZKOR = None if str((values['ДозаK'])) == "" else float(values['ДозаK'])
     ACK = None if str((values['АСК'])) == "" else float(values['АСК'])
     ZAL = None if str((values['Залізо 1 в/м, 2 per os, 3 в/в'])) == "" else float(values['Залізо 1 в/м, 2 per os, 3 в/в'])
     OTER = None if str((values['Оксигенотерапія'])) == "" else float(values['Оксигенотерапія'])
     OPT = None if str((values['Оптимальна Терапія'])) == "" else float(values['Оптимальна Терапія'])
     #14 'Примітки', 'ТЕА 0-рек,1-пров,2-відм.БАП 3-рек,4-пров, 5-відм 6-не рек', 'Дата теа', 'Рекомендации Кулика', 'Декомпенсація', 'Дата дек', 'Трансплантація 1-пров, 0 -рек', 'Дата транс', 'Смерть', 'Дата см'
     PRM = None if str((values['Примітки'])) == "" else str(values['Примітки'])
     TEA = None if str((values['ТЕА 0-рек,1-пров,2-відм.БАП 3-рек,4-пров, 5-відм 6-не рек'])) == "" else float(values['ТЕА 0-рек,1-пров,2-відм.БАП 3-рек,4-пров, 5-відм 6-не рек'])
     DATTEA = None if str((values['Дата теа'])) == "" else str(values['Дата теа'])
     REKK = None if str((values['Рекомендации Кулика'])) == "" else str(values['Рекомендации Кулика'])
     DEK = None if str((values['Декомпенсація'])) == "" else float(values['Декомпенсація'])
     DATADEK = None if str((values['Дата дек'])) == "" else str(values['Дата дек'])
     TRNS = None if str((values['Трансплантація 1-пров, 0 -рек'])) == "" else float(values['Трансплантація 1-пров, 0 -рек'])
     DATATRNS = None if str((values['Дата транс'])) == "" else str(values['Дата транс'])
     SMERT = None if str((values['Смерть'])) == "" else float(values['Смерть'])
     DATASM = None if str((values['Дата см'])) == "" else str(values['Дата см'])

     workbook = load_workbook(filename="patients.xlsx")
     sheet = workbook.active

     if weightp != None and height != None:
        BMI = float(weightp) / ((float(height) / 100 * float(height) / 100))
     else: BMI = None

     data = [Rik,Number,Gos,NoIX,PIB,Per,Birth,Vik,Mis,Pho,Doc,Sex,height,weightp,weightv,BMI,Scarm,Diag,Lg,N11,N12,N13,N141,N142,N143,N144,N145,N1,N2,N3,N41,N42,N5,St1,St2,St3,Fk1,Fk2,Fk3,Fk4,Sn1,Sn2a,Sn2b,Fp,Tp,As,Gip,Plv,Hoz,Soa,ANI,Mins,Gpmk,Anm,Zald,Gp,BBC,Pis,DMP,DSP,Vap,IBBC,Zkd,Tmb,Krv,Sin,Catp,Datp,Csp,Catv,Datv,Csv,Gem,Ert,Gemkr,MCV,MCH,Trobm,Lek,Shoe,MNO,Zag,Tg,K1,Bil,Cre,Cli,Sech,Alt,Ast,Glu,HBsAg,AHCV,Vich,Ntpro1,Ntpro2,TTG,Fer,AENA,Rez,CAT,CRPXM,CRPXE,RCav,LCav,RABI,LABI,xxp1,xxb1,Spo1,Spop1,xxp2,xxb2,Spo2,Spop2,EXOKs,AORT,LP,SLP,INDLP,SPP,INDPP,MSP,ZSLS,KDRLS,KSRLS,KDOLS,KSOLS,YOLS,FV,MMLS,EAMK,DTMK,BSPS,SNPS,SPSD,SPSS,FSPS,POZD,POPR,Ss,TAPSE,SRNT,TASS,LA,NPV,KOLB,SIstl,Sertl,IED,IES,AKN,MKN,TKN,LKN,YLG,VYS,KPS,DATAKAT,STL1,DTL1,SerTLA1,SSTLPS,DSTLPS,SerTPS,SSTLPP,DSTLPP,SerPP,TZLA1,HOK1,SERI1,UV1,UI1,PVR1,PVRw,CSS1,SAT1,DAT1,SerAT1,SVR,TPR,GDD,LVSW,LVSWI,RVSW,RVSWI,VRT,SSTLA2,DSTLA2,SerTLA2,TZLA2,HOK2,SERI2,UV2,UI2,PVR2,CSS2,SAT2,DAT2,SerAT2,Rez2,PROB,STLA,DTLA,SerTLA3,TZLA3,KROVLA,pH1,pO21,pCO21,Hct1,Na1,Ka,Ca1,THb1,HCO31,SvO21,BE1,P50v,Kra,pH2,pO22,pCO22,Hct2,Na2,Kb,Ca2,THb2,HCO32,SvO21,BE2,P50a,Ryz,Spr,Fgel,Fgelp,OFv,OFvp,OFFG,OFFGp,Gel,Gelp,OFG,OFGp,HSK,Do,Ln,VYS2,Bodi,TLS,TLSp,Dif,DLCO,DLCOp,KTMO,Datak,VYS3,Lik,Furv,Furm,Dof,Sli,DOZS,TAD,DOZT,VEN,DOZV,AK,DOZAK,Fyr,DOZF,AA,DOZAA,SERG,DOZSG,VV,DOZVV,ERA,DOZERA,OAK,DOZOAK,TRI,DOZTRI,KOR,DOZKOR,ACK,ZAL,OTER,OPT,PRM,TEA,DATTEA,REKK,DEK,DATADEK,TRNS,DATATRNS,SMERT,DATASM]
     sheet.append(data)
     workbook.save(filename="patients.xlsx")

    elif event == 'clear':
        for item in values:
            window[item].update("")

window.close()
